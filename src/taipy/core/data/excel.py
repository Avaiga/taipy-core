# Copyright 2022 Avaiga Private Limited
#
# Licensed under the Apache License, Version 2.0 (the "License"); you may not use this file except in compliance with
# the License. You may obtain a copy of the License at
#
#        http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software distributed under the License is distributed on
# an "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the License for the
# specific language governing permissions and limitations under the License.

from collections import defaultdict
from datetime import datetime, timedelta
from os.path import isfile
from typing import Any, Dict, List, Optional, Set, Tuple

import pandas as pd
from openpyxl import load_workbook

from taipy.config.common.scope import Scope

from ..common._reload import _self_reload
from ..common.alias import DataNodeId, JobId
from ..exceptions.exceptions import ExposedTypeLengthMismatch, MissingRequiredProperty, NonExistingExcelSheet
from .data_node import DataNode


class ExcelDataNode(DataNode):
    """Data Node stored as an Excel file.

    The Excel file format is _xlsx_.

    Attributes:
        config_id (str): Identifier of this data node configuration. It must be a valid Python
            identifier.
        scope (Scope^): The scope of this data node.
        id (str): The unique identifier of this data node.
        name (str): A user-readable name of this data node.
        parent_id (str): The identifier of the parent (pipeline_id, scenario_id, cycle_id) or
            `None`.
        last_edit_date (datetime): The date and time of the last modification.
        job_ids (List[str]): The ordered list of jobs that have written this data node.
        validity_period (Optional[timedelta]): The validity period of a cacheable data node.
            Implemented as a timedelta. If _validity_period_ is set to None, the data node is
            always up-to-date.
        edit_in_progress (bool): True if a task computing the data node has been submitted
            and not completed yet. False otherwise.
        path (str): The path to the Excel file.
        properties (dict[str, Any]): A dictionary of additional properties. The _properties_
            must have a _"default_path"_ or _"path"_ entry with the path of the Excel file.
    """

    __STORAGE_TYPE = "excel"
    __EXPOSED_TYPE_PROPERTY = "exposed_type"
    __EXPOSED_TYPE_NUMPY = "numpy"
    __PATH_KEY = "path"
    __DEFAULT_PATH_KEY = "default_path"
    __HAS_HEADER_PROPERTY = "has_header"
    __SHEET_NAME_PROPERTY = "sheet_name"
    _REQUIRED_PROPERTIES: List[str] = []

    def __init__(
        self,
        config_id: str,
        scope: Scope,
        id: Optional[DataNodeId] = None,
        name: Optional[str] = None,
        parent_id: Optional[str] = None,
        last_edit_date: Optional[datetime] = None,
        job_ids: List[JobId] = None,
        validity_period: Optional[timedelta] = None,
        edit_in_progress: bool = False,
        properties: Dict = None,
    ):
        if properties is None:
            properties = {}
        if missing := set(self._REQUIRED_PROPERTIES) - set(properties.keys()):
            raise MissingRequiredProperty(
                f"The following properties " f"{', '.join(x for x in missing)} were not informed and are required"
            )

        self._path = properties.get(self.__PATH_KEY, properties.get(self.__DEFAULT_PATH_KEY))
        if self._path is None:
            raise MissingRequiredProperty("default_path is required in a Excel data node config")
        else:
            properties[self.__PATH_KEY] = self._path

        if self.__SHEET_NAME_PROPERTY not in properties.keys():
            properties[self.__SHEET_NAME_PROPERTY] = None
        if self.__HAS_HEADER_PROPERTY not in properties.keys():
            properties[self.__HAS_HEADER_PROPERTY] = True

        super().__init__(
            config_id,
            scope,
            id,
            name,
            parent_id,
            last_edit_date,
            job_ids,
            validity_period,
            edit_in_progress,
            **properties,
        )

        if not self._last_edit_date and isfile(self._path):
            self.unlock_edit()

    @property  # type: ignore
    @_self_reload(DataNode._MANAGER_NAME)
    def path(self):
        return self._path

    @path.setter
    def path(self, value):
        self._path = value
        self.properties[self.__PATH_KEY] = value

    @classmethod
    def storage_type(cls) -> str:
        return cls.__STORAGE_TYPE

    def _read(self):
        if self.__EXPOSED_TYPE_PROPERTY in self.properties:
            if self.properties[self.__EXPOSED_TYPE_PROPERTY] == self.__EXPOSED_TYPE_NUMPY:
                return self._read_as_numpy()
            return self._read_as()
        return self._read_as_pandas_dataframe()

    def __sheet_name_to_list(self, properties):
        if properties[self.__SHEET_NAME_PROPERTY]:
            sheet_names = properties[self.__SHEET_NAME_PROPERTY]
        else:
            excel_file = load_workbook(properties[self.__PATH_KEY])
            sheet_names = excel_file.sheetnames
            excel_file.close()
        return sheet_names if isinstance(sheet_names, (List, Set, Tuple)) else [sheet_names]

    def _read_as(self):
        excel_file = load_workbook(self._path)
        exposed_type = self.properties[self.__EXPOSED_TYPE_PROPERTY]
        work_books = defaultdict()
        sheet_names = excel_file.sheetnames

        if isinstance(exposed_type, dict):
            provided_sheet_names = list(exposed_type.keys())
            for sheet_name in provided_sheet_names:
                if not sheet_name in sheet_names:
                    raise NonExistingExcelSheet(sheet_name, self._path)
            sheet_names = provided_sheet_names
        elif isinstance(exposed_type, List):
            provided_sheet_names = self.__sheet_name_to_list(self.properties)
            if len(provided_sheet_names) != len(self.properties[self.__EXPOSED_TYPE_PROPERTY]):
                raise ExposedTypeLengthMismatch(self.config_id)
            sheet_names = provided_sheet_names
        else:
            # Exposed type is a custom class
            provided_sheet_names = self.__sheet_name_to_list(self.properties)
            for sheet_name in provided_sheet_names:
                if not sheet_name in sheet_names:
                    raise NonExistingExcelSheet(sheet_name, self._path)
            sheet_names = provided_sheet_names

        for i, sheet_name in enumerate(sheet_names):
            work_sheet = excel_file[sheet_name]
            custom_class = exposed_type
            if isinstance(exposed_type, dict):
                custom_class = exposed_type[sheet_name]
            elif isinstance(exposed_type, List):
                custom_class = exposed_type[i]

            res = list()
            for row in work_sheet.rows:
                res.append([col.value for col in row])
            if self.properties[self.__HAS_HEADER_PROPERTY]:
                header = res.pop(0)
                for i, row in enumerate(res):
                    res[i] = custom_class(**dict([[h, r] for h, r in zip(header, row)]))
            else:
                for i, row in enumerate(res):
                    res[i] = custom_class(*row)
            work_books[sheet_name] = res

        excel_file.close()

        if len(sheet_names) == 1:
            return work_books[sheet_names[0]]
        return work_books

    def _read_as_numpy(self):
        sheets = self._read_as_pandas_dataframe()
        if (isinstance(sheets, dict)):
            return {sheet_name: df.to_numpy() for sheet_name, df in sheets.items()}
        return sheets.to_numpy()

    def _read_as_pandas_dataframe(self, usecols: Optional[List[int]] = None, column_names: Optional[List[str]] = None):
        try:
            if self.properties[self.__HAS_HEADER_PROPERTY]:
                if column_names:
                    return pd.read_excel(
                        self._path,
                        sheet_name=self.properties[self.__SHEET_NAME_PROPERTY],
                    )[column_names]
                return pd.read_excel(
                    self._path,
                    sheet_name=self.properties[self.__SHEET_NAME_PROPERTY],
                )
            else:
                if usecols:
                    return pd.read_excel(
                        self._path,
                        header=None,
                        usecols=usecols,
                        sheet_name=self.properties[self.__SHEET_NAME_PROPERTY],
                    )
                return pd.read_excel(
                    self._path,
                    header=None,
                    sheet_name=self.properties[self.__SHEET_NAME_PROPERTY],
                )
        except pd.errors.EmptyDataError:
            return pd.DataFrame()

    def _write(self, data: Any):
        if isinstance(data, Dict) and all([isinstance(x, pd.DataFrame) for x in data.values()]):
            writer = pd.ExcelWriter(self._path)
            for key in data.keys():
                data[key].to_excel(writer, key, index=False)
            writer.save()
        else:
            pd.DataFrame(data).to_excel(self._path, index=False)

    def write_with_column_names(self, data: Any, columns: List[str] = None, job_id: Optional[JobId] = None):
        """Write a set of columns.

        Parameters:
            data (Any): The data to write.
            columns (List[str]): The list of column names to write.
            job_id (JobId^): An optional identifier of the writer.
        """
        if not columns:
            df = pd.DataFrame(data)
        else:
            df = pd.DataFrame(data, columns=columns)
        df.to_excel(self.path, index=False)
        self._last_edit_date = datetime.now()
        if job_id:
            self.job_ids.append(job_id)
